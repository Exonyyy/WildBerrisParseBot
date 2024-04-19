import asyncio
import json
import math

import emoji
import time

from openpyxl import load_workbook
from aiogram import types, Router, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext

from urls import parse_urls
from filters import IsExist, IsProduct, ValidValue
from states import ParseProduct
from keyboards import column_reply_keyboard, inline_keyboard, pagination
from settings import sorts, SHOW_PRODUCT
from parser import parse
from callbacks import UserPage

router = Router()


@router.message(Command("start"))
async def start(message: types.Message, state: FSMContext):
    hello_text = f"{emoji.emojize("\U0001F44B")}Привет!\n" \
                 f"Я бот и люблю собирать информацию о товарах на\n" \
                 f"{emoji.emojize("\U000026A1")}WildBerris{emoji.emojize("\U000026A1")}. Вот что я могу:\n\n" \
                 f"{emoji.emojize("\U00002728")}1. Собрать всю самую важную информацию(цена, размер, цвет, отзывы " \
                 f'и тп) об интересующей вас категории товаров {emoji.emojize("\U00002728")}\n' \
                 f"2. Отобразить её вам. Вы сможете выбрать интересный вам товар и перейти к нему\n\n" \
                 f"{emoji.emojize("\U0001F525") * 3}Жми и получай полезную информацию о любимых товарах "
    await message.answer(hello_text)
    await get_category(message, state)


@router.message(
    IsExist(["Меню", "Главная", "Главное"])
)
@router.callback_query(F.data == "menu")
async def get_category(inp_data: types.Message or types.CallbackQuery, state: FSMContext):
    await state.set_state(ParseProduct.category)
    if type(inp_data) == types.Message:
        await inp_data.answer("Выберите категорию товаров", reply_markup=column_reply_keyboard(
            len(parse_urls.keys()), list(parse_urls.keys())))
    else:
        await inp_data.message.answer("Выберите категорию товаров", reply_markup=column_reply_keyboard(
            len(parse_urls.keys()), list(parse_urls.keys())))


@router.message(
    ParseProduct.category,
    IsExist(parse_urls.keys())
)
async def get_product(message: types.Message, state: FSMContext):
    await state.update_data(category=message.text)
    await state.set_state(ParseProduct.product)
    await message.answer("Выберите тип товара", reply_markup=column_reply_keyboard(
        len(parse_urls[message.text.lower().title()]), list(parse_urls[message.text])))


@router.message(
    ParseProduct.product,
    IsProduct(parse_urls)
)
async def get_sort_type(message: types.Message, state: FSMContext):
    await state.update_data(product=message.text)
    await state.set_state(ParseProduct.sort)
    await message.answer("Выберите тип сортировки", reply_markup=column_reply_keyboard(
        len(sorts.keys()), list(sorts.keys())))


@router.message(
    ParseProduct.sort,
    IsExist(sorts.keys())
)
async def get_product_count(message: types.Message, state: FSMContext):
    await state.update_data(sort=message.text)
    await state.set_state(ParseProduct.count)
    await message.answer("Выберите число товаров от 100 до 10000", reply_markup=column_reply_keyboard(
        10, ['100', '1000', '2000', '3000', '3500', '4000', '5000', '6000', '8000', '100000']
    ))


@router.message(
    ParseProduct.count,
    ValidValue()
)
async def get_user_data(message: types.Message, state: FSMContext):
    await state.update_data(count=int(message.text))
    await message.answer("Cбор информации: 0%", reply_markup=inline_keyboard(1, 1, ["Начать"], ["start_parse"]))


@router.callback_query(F.data == "start_parse")
async def get_data(callback: types.CallbackQuery, state: FSMContext):
    user_data = await state.get_data()
    await asyncio.gather(progress_bar(callback.message, user_data), parse_data(callback.message, user_data))
    await state.clear()


async def progress_bar(message: types.Message, data):
    await message.edit_reply_markup(reply_markup=None)
    needed_time = int(data["count"]) // 100 // 2 if int(data["count"]) > 100 else 1
    last_proc = 0
    for second in range(needed_time + 1):
        time.sleep(1)
        proc = round((second / needed_time) * 100)
        if last_proc < proc:
            if last_proc + 10 < 100:
                last_proc += 10
            await message.edit_text(f"Cбор информации: {proc}%", reply_markup=None)


async def parse_data(message: types.Message, data):
    loop = asyncio.get_event_loop()
    collect_data = loop.run_in_executor(
        None, parse, parse_urls[data['category']][data["product"]], data['sort'], int(data['count']) // 100)
    await collect_data
    await message.answer("Сбор данных завершёт", reply_markup=inline_keyboard(
        2, 2, ["Сохранить в таблице", "Показать товары"], ['save_table', 'show_data']))


@router.callback_query(F.data == "show_data")
async def show_data(callback: types.CallbackQuery):
    with open("bot/files/result.json", "r") as data_file:
        collected_data = json.load(data_file)
    pages = math.ceil(len(collected_data)/SHOW_PRODUCT)
    parsed_data = ""
    for product in range(1, SHOW_PRODUCT+1):
        product = str(product)
        parsed_data += f'{collected_data[product]['name']}\n' \
                       f'Цена: {collected_data[product]['cost']}руб / {collected_data[product]['wallet_cost']}' \
                       f'руб (с кошельком)\n' \
                       f'Оценка: {collected_data[product]['rate']}/5, {collected_data[product]['feedback']} отзывов\n' \
                       f'Цвета: {", ".join(collected_data[product]['colors'])}\n' \
                       f'Размеры: {", ".join(collected_data[product]["sizes"])}\n' \
                       f'https://www.wildberries.ru/catalog/{collected_data[product]["article"]}/detail.aspx' \
                       f'?targetUrl=SP\n\n'
    await callback.message.edit_text(parsed_data, reply_markup=pagination(pages))


@router.callback_query(UserPage.filter(F.event == "next"))
async def next_page(callback: types.CallbackQuery, callback_data: UserPage):
    with open("bot/files/result.json", "r") as data_file:
        collected_data = json.load(data_file)
    pages = math.ceil(len(collected_data)/SHOW_PRODUCT)
    page = callback_data.page + 1 if callback_data.page < pages else callback_data.page
    parsed_data = ''
    last_product = page*SHOW_PRODUCT+1 if page*SHOW_PRODUCT < len(collected_data) else len(collected_data)+1
    for product in range((page-1)*SHOW_PRODUCT+1, last_product):
        product = str(product)
        parsed_data += f'{collected_data[product]['name']}\n' \
                       f'Цена: {collected_data[product]['cost']}руб / {collected_data[product]['wallet_cost']}' \
                       f'руб (с кошельком)\n' \
                       f'Оценка: {collected_data[product]['rate']}/5, {collected_data[product]['feedback']} отзывов\n' \
                       f'Цвета: {", ".join(collected_data[product]['colors'])}\n' \
                       f'Размеры: {", ".join(collected_data[product]["sizes"])}\n' \
                       f'https://www.wildberries.ru/catalog/{collected_data[product]["article"]}/detail.aspx' \
                       f'?targetUrl=SP\n\n'
    if callback_data.page != page:
        await callback.message.edit_text(parsed_data, reply_markup=pagination(pages, page))


@router.callback_query(UserPage.filter(F.event == "prev"))
async def prev_page(callback: types.CallbackQuery, callback_data: UserPage):
    with open("bot/files/result.json", "r") as data_file:
        collected_data = json.load(data_file)
    pages = math.ceil(len(collected_data)/SHOW_PRODUCT)
    page = callback_data.page - 1 if callback_data.page > 1 else 1
    parsed_data = ''
    for product in range((page-1)*SHOW_PRODUCT+1, page*SHOW_PRODUCT+1):
        product = str(product)
        parsed_data += f'{collected_data[product]['name']}\n' \
                       f'Цена: {collected_data[product]['cost']}руб / {collected_data[product]['wallet_cost']}' \
                       f'руб (с кошельком)\n' \
                       f'Оценка: {collected_data[product]['rate']}/5, {collected_data[product]['feedback']} отзывов\n' \
                       f'Цвета: {", ".join(collected_data[product]['colors'])}\n' \
                       f'Размеры: {", ".join(collected_data[product]["sizes"])}\n' \
                       f'https://www.wildberries.ru/catalog/{collected_data[product]["article"]}/detail.aspx' \
                       f'?targetUrl=SP\n\n'
    if page != callback_data.page:
        await callback.message.edit_text(parsed_data, reply_markup=pagination(pages, page))


@router.callback_query(F.data == "save_table")
async def save_table(callback: types.CallbackQuery):
    await callback.message.edit_reply_markup(reply_markup=None)
    file_name = "bot/files/result.xlsx"
    tables = load_workbook(file_name)
    table = tables["result_data"]

    with open("bot/files/result.json", "r") as data_file:
        collected_data = json.load(data_file)

    column_names = ["Название", "Цена", "Цена с кошельком", "Рейтинг", "Количество отзывов", "Цвета", "Размеры",
                    "Артикль"]
    json_names = ["name", "cost", "wallet_cost", "rate", "feedback", "colors", "sizes", "article"]
    for column in range(1, len(column_names)+1):
        cell = table.cell(row=1, column=column)
        cell.value = column_names[column-1]

    for row in range(2, len(collected_data)+1):

        for column in range(1, len(collected_data[str(row)])+1):
            if type(collected_data[str(row)][json_names[column-1]]) == int:
                value = str(collected_data[str(row)][json_names[column-1]])
            elif type(collected_data[str(row)][json_names[column-1]]) == list:
                value = ", ".join(collected_data[str(row)][json_names[column-1]])
            else:
                value = collected_data[str(row)][json_names[column-1]]
            cell = table.cell(row=row, column=column)
            cell.value = value

    tables.save(file_name)
    tables.close()
    await callback.message.answer_document(types.input_file.FSInputFile("bot/files/result.xlsx"),
                                           reply_markup=inline_keyboard(1, 1, ["Меню"], ["menu"]))
